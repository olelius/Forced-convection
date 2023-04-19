import os
# import main
from PyQt5.QtWidgets import QDialog, QMessageBox, QApplication, QTableWidget, QFileDialog, QHeaderView, QPushButton, QTableWidgetItem
from PyQt5.QtCore import Qt, QTimer, QEvent
from PyQt5.QtGui import QIcon
from openpyxl import Workbook


class NewWindow(QDialog):
    def __init__(self, conn, main_window):

        super().__init__()

        # 创建应用程序图标对象
        app_icon = QIcon("logo.jpg")

        # 设置应用程序图标
        self.setWindowIcon(app_icon)

        self.main_window = main_window

        self.setWindowTitle("数据表格")

        # 连接数据库
        self.conn = conn
        self.c = self.conn.cursor()
        
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_table)
        self.timer.start(100)
        self.timer.stop()

        # 安装事件过滤器
        self.installEventFilter(self)

        # 创建表格
        self.table = QTableWidget(self)
        self.table.setGeometry(20, 50, 913, 410)
        self.table.verticalHeader().setDefaultSectionSize(10)  # 添加垂直代码
        self.table.setRowCount(10)
        self.table.setColumnCount(19)
        self.table.setHorizontalHeaderLabels(
            ["功率", "风压", "通道1", "通道2", "通道3", "通道4", "通道5", "通道6", "通道7",
             "通道8", "通道9", "通道10", "通道11", "通道12", "通道13", "通道14", "通道15", "通道16", "来流"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setShowGrid(True)
        self.table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        # self.table.verticalHeader().setStyleSheet("border: 1px solid;")
        # self.table.horizontalHeader().setStyleSheet("border: 1px solid;")



        for i in range(20):
            self.table.setColumnWidth(i, 47)
        for i in range(self.table.rowCount()):
            self.table.setRowHeight(i, 36)

        # 禁止调整行列大小
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)

        # 添加表格内容
        self.update_table()

        # 设置窗口居中显示
        self.setFixedSize(960, 540)     # 固定窗口大小
        screen = QApplication.desktop().screenGeometry()
        self.setGeometry(screen.width() // 3, screen.height() // 4, 960, 540)

        # 添加刷新按钮
        self.refresh_button = QPushButton("刷新", self)
        self.refresh_button.setGeometry(750, 470, 80, 40)
        self.refresh_button.clicked.connect(self.refresh_table)

        # 添加保存按钮
        self.save_button = QPushButton("保存", self)
        self.save_button.setGeometry(860, 470, 80, 40)
        self.save_button.clicked.connect(self.save_table)

    def update_table(self):
        # 清空表格
        self.table.clearContents()

        # 从数据库中读取数据
        self.c.execute("SELECT * FROM data")
        rows = self.c.fetchall()

        # 更新表格内容
        for i in range(len(rows)):
            row = rows[i]
            for j in range(3, 21):
                if j == 3:
                    item = QTableWidgetItem("{:.2f}".format(row[j])+'W')
                elif j == 4:
                    item = QTableWidgetItem("{:.0f}".format(row[j])+'Pa')
                else:
                    item = QTableWidgetItem("{:.1f}".format(row[j])+'℃')
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(i, j-3, item)

    def save_table(self):
        # 打开文件对话框，选择保存文件的位置和文件名
        options = QFileDialog.Options()
        # options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getSaveFileName(
            self, "保存文件", f"{os.path.expanduser('~')}/Desktop/", "Excel Files (*.xlsx);;文本文件 (*.txt);;所有文件 (*)", options=options)

        # 如果有选择文件名，则保存数据到文件中
        if file_name:
            if file_name.endswith(".xlsx"):  # 保存为Excel文件
                wb = Workbook()
                ws = wb.active

                # 写入表头
                for col in range(self.table.columnCount()):
                    ws.cell(row=1, column=col+1,
                            value=self.table.horizontalHeaderItem(col).text())

                # 写入数据
                for row in range(self.table.rowCount()):
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        if item is not None:
                            ws.cell(row=row+2, column=col+1,
                                    value=item.text())

                wb.save(file_name)
            else:  # 保存为文本文件
                with open(file_name, "w", encoding='utf-8') as file:
                    # 写入表头
                    header = "\t".join(self.table.horizontalHeaderItem(i).text()
                                       for i in range(self.table.columnCount()))
                    file.write(header + "\n")

                    # 写入数据
                    for row in range(self.table.rowCount()):
                        line = ""
                        for col in range(self.table.columnCount()):
                            item = self.table.item(row, col)
                            if item is not None and item.text():
                                line += item.text()
                            if col < self.table.columnCount() - 1:
                                line += "\t"
                        file.write(line + "\n")

            print("保存成功")   # 可以弹出对话框提示保存成功
            # 弹出提示窗口
            QMessageBox.information(self, "提示", "文件已成功保存！", QMessageBox.Ok)

    # 刷新按钮
    def refresh_table(self):
        # 更新表格
        self.update_table()

    # 拖动时禁止主界面数据更新，解决卡顿问题
    def read_main_button_text(self):
        # 从 main_window 中的 button 读取文本值
        button_text = self.main_window.button.text()
        return button_text
    
    def eventFilter(self, source, event):
        if  event.type() == QEvent.NonClientAreaMouseButtonPress:
            # 停止主窗口中的计时器
                self.main_window.timer.stop()
        elif  event.type() == QEvent.NonClientAreaMouseButtonRelease:
            # 启动主窗口中的计时器
            if self.read_main_button_text() == "暂停":
                self.main_window.timer.start(100)
        return super().eventFilter(source, event)

    def closeEvent(self, event):
        # 关闭窗口时关闭数据库连接和定时器
        event.accept()
